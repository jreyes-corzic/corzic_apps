import odoo_access as od
import pandas as pd
import math
import os, shutil, zipfile
import openpyxl as opxl

from datetime import datetime
from pathlib import Path

from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl import load_workbook, Workbook

import requests



class ariba_consts:
	SUPPLIER_ID='an11091553062'
	UOM='EA'
	TEMPLATE_FILE=str(Path(__file__).with_name('SUFS_Import_Template.xlsx'))
	GUITAR_SERVICE_COST = 75.0
	MANDOLIN_SERVICE_COST = 55.0
	BASS_SERVICE_COST = 95.0
	
def get_odoo_connection():
	uid,models=od.access_odoo_db(od.url,od.db,od.username,od.password)
	print(f"retrieved odoo access with id: {uid}")
	return uid,models

	
def get_output_path():
	default_path=os.path.expanduser('~/Library/CloudStorage/OneDrive-Corzic,LLC/Corzic, LLC - Documents/Products/SUFS_Catalogs/')
	output_path=os.getenv('SUFS_OUTPUT_DIR')
	if not output_path:
		output_path=default_path if os.path.isdir(default_path) else os.path.join(os.getcwd(),'output')
	Path(output_path).mkdir(parents=True,exist_ok=True)
	return output_path



def get_df_upload(db,uid,password,models,override=None):
	override=set(override or [])
	sufs_ids=od.get_ids(db,uid,password,models,od.models[od.models.index('product.template')],[[('sales_channel_ids','=',[3]),('sufs_status','=','listed')]])
	sufs_list=od.read_ids(db,uid,password,models,od.models[od.models.index('product.template')],sufs_ids,['id','sufs_name','default_code','product_brand_id','sufs_cost','sufs_description','spsc_code','barcode','sufs_lead_time','supplier_url','manufacturer_url','sufs_image_url','sufs_thumbnail_url','product_tag_ids','ovap_onhand','quarantine_onhand','qty_available','incoming_qty','outgoing_qty','route_ids'])

	
	sufs_df=pd.DataFrame(sufs_list)
	expected_skus=set(sufs_df['default_code'].dropna())
	unknown_override_skus=sorted(override - expected_skus)
	
	ids_to_remove=[]
	skus_to_remove=[]
	for idx,row in sufs_df.iterrows():
		if (row['qty_available'] - row['outgoing_qty'] - row['ovap_onhand'] < 1):
			if (6 not in row['route_ids'] and not row['default_code'] in override):
				ids_to_remove.append(row['id'])
				skus_to_remove.append(row['default_code'])
			
	sufs_df=sufs_df[~sufs_df['id'].isin(ids_to_remove)]
	sufs_df.reset_index(drop=True)
	return sufs_df,len(ids_to_remove),skus_to_remove,unknown_override_skus

	

def generate_output_df(sufs_df):
		
		
	
	item_count=len(sufs_df)
	supplier_ids = [ariba_consts.SUPPLIER_ID]*item_count
	part_ids = sufs_df['default_code']
	manufacturer_ids = sufs_df['barcode']
	descriptions = sufs_df['sufs_description']
	spsc = sufs_df['spsc_code']
	
	for idx,row in sufs_df.iterrows():
		old_cost=row['sufs_cost']
		#print(row['sufs_cost'])
		if (2 in row['product_tag_ids']):
			sufs_df.at[idx,'sufs_cost'] = old_cost + ariba_consts.GUITAR_SERVICE_COST
			#print(f'adding guitar plek cost: ${ariba_consts.GUITAR_SERVICE_COST}')
		elif (4 in row['product_tag_ids']):
			sufs_df.at[idx,'sufs_cost'] =old_cost + ariba_consts.MANDOLIN_SERVICE_COST
			#print(f'adding mandolin plek cost')
		elif (5 in row['product_tag_ids']):
			sufs_df.at[idx,'sufs_cost'] =old_cost + ariba_consts.BASS_SERVICE_COST
			#print(f'adding bass plek cost')
		#print(sufs_df.at[idx,'sufs_cost'])
		#print(f'Updated cost from ${old_cost} to ${sufs_df.at[idx,'sufs_cost']}')
	
	prices=sufs_df['sufs_cost']
	uom = [ariba_consts.UOM]*item_count
	lead_times = sufs_df['sufs_lead_time']

	for idx,row in sufs_df.iterrows():
		vals=row['product_brand_id']
		if (not vals):
			sufs_df.at[idx,'product_brand_id']=None
		else:
			sufs_df.at[idx,'product_brand_id'] = vals[1]

	brands = sufs_df['product_brand_id']
	supplier_urls=sufs_df['supplier_url']
	manufacturer_urls=sufs_df['manufacturer_url']
	market_price=['']*item_count
	short_names=sufs_df['sufs_name']
	images=sufs_df['sufs_image_url']
	thumbs=sufs_df['sufs_thumbnail_url']

	#export dataframe for excel sheet 2
	data=list(zip(supplier_ids,part_ids,manufacturer_ids,descriptions,spsc,prices,uom,lead_times,brands,supplier_urls,manufacturer_urls,market_price,short_names,images,thumbs))
	columns=['Supplier ID','Supplier Part ID','Manufacturer Part ID','Item Description','SPSC Code','Unit Price','Unit of Measure','Lead Time','Manufacturer Name','Supplier URL','Manufacturer URL','Market Price','Short Name','Image','Thumbnail']
	out_df= pd.DataFrame(data,columns=columns)
		
	
	return out_df



def generate_export(sufs_df,removal_list,db,uid,password,models,out_type='upload'):
	time_stamp=datetime.now().strftime("%m/%d/%Y")
	file_time_stamp=datetime.now().strftime("%y%m%d")
	item_count=len(sufs_df)

	template_wb = load_workbook(ariba_consts.TEMPLATE_FILE)
	ws1=template_wb['Sheet1']

	#prepare the export workbook header values to include the correct item count and the correct timestamp 
	count_row=0
	time_row=0
	data_start=0
	for row_idx in range(1,ws1.max_row+1):
		label_cell=ws1.cell(row=row_idx,column=1)
		#print(label_cell.value)
		if (label_cell.value == 'ITEMCOUNT:'):
			count_row=row_idx
			#print(f'Found item count row at {count_row}')
		if (label_cell.value == 'TIMESTAMP:'):
			time_row=row_idx
			#print(f'Found time stamp row at {time_row}')
		if (label_cell.value == 'DATA'):
			data_start=row_idx+1
			#print(f'Found data start row at {data_start}')

	ws1.cell(row=count_row,column=2).value = item_count
	ws1.cell(row=time_row,column=2).value = str(time_stamp)

	#save template file as ariba export file for formatted uploads
	ariba_file_name=f'{file_time_stamp} -  Corzic PRD Import - {out_type}'
	template_wb.save(ariba_file_name+'.xlsx')

	

	out_filename=ariba_file_name+'.xlsx'
	with pd.ExcelWriter(out_filename, mode='a', engine='openpyxl') as writer:
   		sufs_df.to_excel(writer, sheet_name='Sheet2', index=False)
    

	#move sheet 2 to main sheet 
	wb=load_workbook(out_filename)
	ws1=wb['Sheet1']
	ws2=wb['Sheet2']

	#format cells for Ariba upload 
	for row_idx in range(2,ws2.max_row+1):
		row_data=[]
		for col_idx in range(1,ws2.max_column+1):
			cell_value=ws2.cell(row=row_idx,column=col_idx).value
			if (cell_value == False):
				row_data.append(None)
			else:
				row_data.append(cell_value)
		ws1.append(row_data)

	ws1.cell(row=ws1.max_row+1,column=1).value='ENDOFDATA'

	for row_idx in range(data_start,ws1.max_row+1):
		if (ws1[f'C{row_idx}'].value != None):		
			ws1[f'C{row_idx}'].value = int(ws1[f'C{row_idx}'].value)
			ws1[f'C{row_idx}'].number_format='0'
	
		if (ws1[f'E{row_idx}'].value != None):
			ws1[f'E{row_idx}'].value=int(ws1[f'E{row_idx}'].value)
			ws1[f'E{row_idx}'].number_format='0'

	#remove sheet 2 and save
	wb.remove(wb['Sheet2'])
	output_path=get_output_path()
	out_path=os.path.join(output_path,out_filename)
	wb.save(out_path)
	os.remove(out_filename)
	
	#removal list excel document
	removal_data={'sku_removed':removal_list}
	removal_df=pd.DataFrame(removal_data)
	
	#update odoo sufs status to 'stock out'
	for idx,row in removal_df.iterrows():
		prod_id=od.get_ids(db,uid,password,models,od.models[od.models.index('product.template')],[[('default_code','=',row['sku_removed'])]])
		field={'sufs_status':'stock_out'}
		ref=od.update_id(db,uid,password,models,od.models[od.models.index('product.template')],prod_id[0],field)
	list_filename=f'SKUs Removed From Catalog - {file_time_stamp}.xlsx'
	list_path=os.path.join(output_path,list_filename)
	removal_df.to_excel(list_path,index=False)
	return out_path,list_path


def get_catalog_upload(override_list=None):
	uid,models=get_odoo_connection()
	db=od.db
	password=od.password
	sufs_df,num_removals,list_removals,unknown_override_skus=get_df_upload(db,uid,password,models,override_list)
	default_out=generate_output_df(sufs_df)
	

	return generate_export(default_out,list_removals,db,uid,password,models),num_removals,unknown_override_skus


def zip_catalogs(exports):

	ariba_zip="Ariba-Import-Files.zip"
	with zipfile.ZipFile(ariba_zip,'w',zipfile.ZIP_DEFLATED) as zf:
		for export in exports:
			zf.write(export,arcname=os.path.basename(export))

	print(f"Ariba Import Files created and compressed into {ariba_zip}")
	for export in exports:
		os.remove(export)
	
	return ariba_zip

if __name__ == "__main__":
	exports,num_removals,unknown_override_skus=get_catalog_upload()
	print(f"Generated {len(exports)} files. Removed {num_removals} SKUs.")
	for export in exports:
		print(export)
	if unknown_override_skus:
		print(f"Unknown override SKUs: {', '.join(unknown_override_skus)}")
